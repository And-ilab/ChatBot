import logging
import pytz
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import get_user_model
from django.forms import CharField
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Exists, OuterRef, Subquery, Value, Case, When, F, Count
from django.db.models.functions import TruncDate, Concat
from django.utils.timezone import now
from .models import Dialog, Message, User, TrainingMessage, Settings, Documents, ActionLog
from django.db.models import CharField
from chat_user.models import ChatUser, Session
from .forms import UserForm, UserFormUpdate
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import json
import re
import spacy
import os
import os
from django.conf import settings
import pymorphy3
from config import config_settings
from django.conf import settings
import requests
from django.db.models import Q
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta, datetime
from chat_user.models import ChatUser
from django.core.mail import send_mail
from authentication.decorators import role_required
import random
import uuid

logger = logging.getLogger('chat_dashboard')
user_action = logging.getLogger('user_actions')


def log_action(action_text):
    try:
        print('SUKA')
        # if not user or not isinstance(user, User):
        #     return False
        #
        # if not action_text or not isinstance(action_text, str):
        #     return False

        # role_display = user.get_role_display()
        # last_name = user.last_name if user.last_name else ""
        formatted_action = f"{'Администратор'} -- {action_text}"
        print(formatted_action)

        ActionLog.objects.create(user=None, action=formatted_action)
        return True

    except Exception:
        return False

#@role_required(['admin', 'operator'])
def analytics(request):
    user = request.user
    # user_action.info(
    #     'Accessing analytics page',
    #     extra={
    #         'user_id': user.id,
    #         'user_name': user.first_name + ' ' + user.last_name,
    #         'action_type': 'Accessing analytics page',
    #         'time': datetime.now(),
    #         'details': json.dumps({
    #             'status': f"{user.first_name} {user.last_name}' get access to analytics",
    #         })
    #
    #     }
    # )
    """Displays the analytics page."""
    logger.info("Accessing analytics page.")
    return render(request, 'chat_dashboard/analytics.html')

#@role_required(['admin', 'operator'])
def logs_view(request):
    logger.info("Accessing logs page.")

    period = int(request.GET.get('period', 0))
    user_id = request.GET.get('user_id', '').strip()

    logs = ActionLog.objects.all()

    # Фильтрация по дате
    if period == 1:  # Сегодня
        today = datetime.now().date()
        logs = logs.filter(timestamp__date=today)
    elif period == 7:  # За последние 7 дней
        start_date = datetime.now() - timedelta(days=7)
        logs = logs.filter(timestamp__gte=start_date)
    elif period == 30:  # За последние 30 дней
        start_date = datetime.now() - timedelta(days=30)
        logs = logs.filter(timestamp__gte=start_date)

    # Фильтрация по ID пользователя
    if user_id.isdigit():
        logs = logs.filter(user_id=user_id)

    logs = logs.order_by('-timestamp')

    # Форматирование логов
    moscow_tz = pytz.timezone('Europe/Moscow')
    formatted_logs = []
    for log in logs:
        local_time = log.timestamp.astimezone(moscow_tz)
        formatted_time = local_time.strftime('%Y-%m-%d %H:%M:%S')
        formatted_logs.append(f"[{formatted_time}] - {log.action}")

    logs_text = "\n".join(formatted_logs)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'logs_text': logs_text})

    return render(request, 'chat_dashboard/logs.html', {'logs_text': logs_text})

@csrf_exempt
def send_message(request, dialog_id):
    """Sends a message in the specified dialog."""
    logger.info(f"Sending message to dialog ID: {dialog_id}")
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            content = data.get('content')
            sender_type = data.get('sender_type')
            sender_id = data.get('sender_id')
            timestamp = data.get('timestamp')
            message_type = data.get('message_type')

            logger.debug(
                f"Message data: content={content}, sender_type={sender_type}, sender_id={sender_id}, timestamp={timestamp}")

            if content and sender_type and timestamp and message_type:
                dialog = Dialog.objects.get(id=dialog_id)
                Message.objects.create(
                    dialog=dialog,
                    sender_type=sender_type,
                    sender_id=sender_id if sender_type == 'user' else None,
                    message_type=message_type,
                    content=content,
                    created_at=timestamp
                )
                logger.info("Message sent successfully.")
                return JsonResponse({'status': 'success', 'message': 'Message sent'})
            logger.warning("Invalid data: content or sender_type is missing.")
            return JsonResponse({'status': 'error', 'message': 'Invalid data'}, status=400)
        except Exception as e:
            logger.exception("An error occurred while sending a message.")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    logger.warning("Invalid method: only POST is supported.")
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


def messages_count_data(request):
    """Returns message counts grouped by date and sender type."""
    logger.info("Fetching messages count data.")
    data = (
        Message.objects.annotate(message_date=TruncDate('created_at'))
        .values('message_date', 'sender_type')
        .annotate(count=Count('id'))
        .order_by('message_date', 'sender_type')
    )

    formatted_data = {}
    for entry in data:
        date = entry['message_date']
        sender_type = entry['sender_type']
        count = entry['count']
        if date not in formatted_data:
            formatted_data[date] = {'user': 0, 'bot': 0}
        formatted_data[date][sender_type] = count

    response_data = [
        {'created_at': date, 'user': values['user'], 'bot': values['bot']}
        for date, values in formatted_data.items()
    ]
    logger.debug(f"Formatted message count data: {response_data}")
    return JsonResponse(response_data, safe=False)


def daily_messages_data(request):
    """Returns daily training messages count."""
    logger.info("Fetching daily messages data.")
    data = (
        TrainingMessage.objects
        .annotate(date=TruncDate('created_at'))
        .values('date')
        .annotate(count=Count('id'))
        .order_by('date')
    )
    logger.debug(f"Daily messages data retrieved: {list(data)}")
    return JsonResponse(list(data), safe=False)


nlp = spacy.load("ru_core_news_sm")
morph = pymorphy3.MorphAnalyzer()
custom_stop_words = {"может", "могут", "какой", "какая", "какое", "какие", "что", "кто", "где", "когда", "зачем",
                     "почему"}


#@role_required(['admin', 'operator'])
def training_dashboard(request):
    """Displays the training dashboard."""
    logger.info("Accessing training dashboard.")
    user = request.user
    unread_messages = TrainingMessage.objects.filter(is_unread=True, is_ignored=False).values(
        'id', 'content', 'created_at'
    )
    ignored_messages = TrainingMessage.objects.filter(is_ignored=True).values(
        'id', 'content', 'created_at'
    )

    context = {
        'all_messages': list(unread_messages) + list(ignored_messages),
        'unread_messages': list(unread_messages),
        'ignored_messages': list(ignored_messages),
    }
    # user_action.info(
    #     'Accessing training dashboard',
    #     extra={
    #         'user_id': user.id,
    #         'user_name': user.first_name + ' ' + user.last_name,
    #         'action_type': 'Accessing training dashboard',
    #         'time': datetime.now(),
    #         'details': json.dumps({
    #             'status': f"{user.first_name} {user.last_name}' get access to training dashboard",
    #         })
    #
    #     }
    # )
    logger.debug(f"Unread messages: {list(unread_messages)}, Ignored messages: {list(ignored_messages)}")
    return render(request, 'chat_dashboard/training.html', context)




@csrf_exempt
def mark_question_trained(request):
    try:
        user = request.user
        data = json.loads(request.body)
        message_id = data.get('message_id')
        sender = data.get('sender_id')
        answer = data.get('answer')

        if not message_id:
            return JsonResponse({'error': 'Не передан идентификатор сообщения.'}, status=400)

        training_message = TrainingMessage.objects.get(id=message_id)
        sender = ChatUser.objects.get(id=sender)
        last_dialog = Dialog.objects.filter(user=sender).order_by('-started_at').first()

        greetings = ["Привет! Спасибо за ваше терпение. Вот ответ на ваш вопрос от нашего оператора. Надеюсь, он будет полезен! 😊",
                     "Здравствуйте! Ваш вопрос обработан, и вот ответ от нашего специалиста. Если что-то непонятно, дайте знать — я всегда готов помочь! 🚀",
                     "Приветствую! Вот ответ на ваш вопрос. Спасибо, что обратились к нам. Если остались вопросы, задавайте — я здесь, чтобы помочь! 📩",
                     "Добрый день! Ваш вопрос рассмотрен, и вот ответ от оператора. Надеюсь, он решит вашу задачу. Если что-то ещё нужно, просто спросите! 😄",
                     "Привет! Вот ответ на ваш вопрос. Спасибо за обращение! Если нужно что-то уточнить, я всегда на связи. 🛎️",
                     "Здравствуйте! Ваш вопрос обработан, и вот ответ от нашего специалиста. Надеюсь, он будет полезен. Если есть ещё вопросы, задавайте! 📨",
                     "Привет! Вот ответ на ваш вопрос. Спасибо за терпение! Если что-то непонятно или нужно уточнить, дайте знать — я готов помочь. 🕒",
                     "Добрый день! Ваш вопрос рассмотрен, и вот ответ от оператора. Надеюсь, он решит вашу задачу. Если что-то ещё нужно, просто спросите! 😊",
                     "Приветствую! Вот ответ на ваш вопрос. Спасибо за обращение! Если нужно что-то уточнить, я всегда на связи. 🚀",
                     "Здравствуйте! Ваш вопрос обработан, и вот ответ от нашего специалиста. Надеюсь, он будет полезен. Если есть ещё вопросы, задавайте! 📩",]
        random_greeting = random.choice(greetings)

        if last_dialog is None:
            return JsonResponse({'error': 'У пользователя нет активных диалогов.'}, status=400)

        Message.objects.create(
            dialog=last_dialog,
            sender_type='bot',
            sender=sender,
            content=f'''{random_greeting}<br>
                        Ваш вопрос:<br>
                        {training_message.content}<br>
                        <br>
                        Ответ оператора:<br>
                        {answer}''',
            message_type='message'
        )

        if training_message.sender and training_message.sender.email:
            subject = "Ваш вопрос обработан — ответ ждёт вас в чате"
            message = (
                f''' Уважаемый(ая) {sender.first_name} {sender.last_name}!,
                Большое спасибо за ваше терпение! Мы рады сообщить, что ваш вопрос был обработан, и ответ уже направлен вам в чат.
                Для вашего удобства дублируем вопрос и ответ ниже:
                Ваш вопрос:
                {training_message.content}

                Ответ оператора:
                {answer}

                Если у вас остались вопросы или что-то требует уточнения, пожалуйста, напишите нам в чат — мы всегда готовы помочь!

                С уважением,
                Команда поддержки Интеллектуальной платформы взаимодействия с пользователями HelpDeskBot'''
             )
            from_email = 'sapunowdany@yandex.by'
            send_mail(subject, message, from_email, [training_message.sender.email],fail_silently=False)
        training_message.delete()
        return JsonResponse({'status': 'success'})
    except TrainingMessage.DoesNotExist:
        # user_action.error(
        #     'Mark_question_trained_unsuccessfully',
        #     extra={
        #         'user_id': user.id,
        #         'user_name': user.first_name + ' ' + user.last_name,
        #         'action_type': 'mark question trained',
        #         'time': datetime.now(),
        #         'details': json.dumps({
        #             'status': f"Message not found",
        #         })
        #
        #     }
        # )
        return JsonResponse({'error': 'Сообщение для дообучения не найдено.'}, status=404)
    except Exception as e:
        # user_action.error(
        #     'Mark_question_trained_unsuccessfully',
        #     extra={
        #         'user_id': user.id,
        #         'user_name': user.first_name + ' ' + user.last_name,
        #         'action_type': 'mark question trained',
        #         'time': datetime.now(),
        #         'details': json.dumps({
        #             'status': f"Error {e}",
        #         })
        #
        #     }
        # )
        return JsonResponse({'error': str(e)}, status=500)


#@role_required(['admin', 'operator'])
def train_message(request, message_id):
    """Displays a message for training."""
    logger.info(f"Accessing training page for message ID: {message_id}")
    user_message = get_object_or_404(TrainingMessage, id=message_id)
    logger.debug(f"Training message retrieved: {user_message.content}")
    user = request.user
    # user_action.info(
    #     'f"Accessing training page for message ID: {message_id}"',
    #     extra={
    #         'user_id': user.id,
    #         'user_name': user.first_name + ' ' + user.last_name,
    #         'action_type': 'accessing train_message',
    #         'time': datetime.now(),
    #         'details': json.dumps({
    #             'status': f"{user.first_name} {user.last_name}' get accessing training page for message ID: {message_id}",
    #         })
    #
    #     }
    # )
    return render(request, 'chat_dashboard/train_message.html',
                  {'user_message': user_message})


def ignore_message(request, message_id):
    """Toggle the ignored status of a training message."""
    logger.info(f"Toggling ignore status for message ID: {message_id}")
    user = request.user
    try:
        message = TrainingMessage.objects.get(id=message_id)
        message.is_unread = False
        message.is_ignored = True
        message.save()

        unread_count = TrainingMessage.objects.filter(is_unread=True, is_ignored=False).count()
        ignored_count = TrainingMessage.objects.filter(is_ignored=True).count()

        logger.info(f"Message {message_id} updated: unread={message.is_unread}, ignored={message.is_ignored}")
        # user_action.info(
        #     'f"toggle_ignore_message successfully"',
        #     extra={
        #         'user_id': user.id,
        #         'user_name': user.first_name + ' ' + user.last_name,
        #         'action_type': 'Toggle ignore message',
        #         'time': datetime.now(),
        #         'details': json.dumps({
        #             'status': f"Toggle ignore message is successful",
        #         })
        #
        #     }
        # )
        return JsonResponse({
            'unread_count': unread_count,
            'ignored_count': ignored_count,
            'message_id': message.id,
            'is_unread': message.is_unread,
            'is_ignored': message.is_ignored,
        })
    except TrainingMessage.DoesNotExist:
        # user_action.error(
        #     'f"TrainingMessage is not found"',
        #     extra={
        #         'user_id': user.id,
        #         'user_name': user.first_name + ' ' + user.last_name,
        #         'action_type': 'TrainingMessage is not found',
        #         'time': datetime.now(),
        #         'details': json.dumps({
        #             'status': f"TrainingMessage is not found",
        #         })
        #     }
        # )
        logger.error(f"Message ID {message_id} not found.")
        return JsonResponse({'error': 'Message not found'}, status=404)


def delete_training_message(request, message_id):
    """Deletes a training message."""
    logger.info(f"Attempting to delete message ID: {message_id}")
    try:
        message = TrainingMessage.objects.get(id=message_id)
        message.delete()
        user = request.user
        unread_count = TrainingMessage.objects.filter(is_ignored=False).count()
        ignored_count = TrainingMessage.objects.filter(is_ignored=True).count()

        logger.info(f"Message ID {message_id} deleted successfully.")
        # user_action.info(
        #     f"Message ID {message_id} deleted successfully.",
        #     extra={
        #         'user_id': user.id,
        #         'user_name': user.first_name + ' ' + user.last_name,
        #         'action_type': 'delete message',
        #         'time': datetime.now(),
        #         'details': json.dumps({
        #             'status': f"{user.first_name} {user.last_name}' delete message with id = {message_id} successfully.",
        #         })
        #
        #     }
        # )
        return JsonResponse({
            'unread_count': unread_count,
            'ignored_count': ignored_count
        })

    except TrainingMessage.DoesNotExist:
        # user_action.error(
        #     f"Message not found",
        #     extra={
        #         'user_id': user.id,
        #         'user_name': user.first_name + ' ' + user.last_name,
        #         'action_type': 'Message not found',
        #         'time': datetime.now(),
        #         'details': json.dumps({
        #             'status': f"Message not found",
        #         })
        #
        #     }
        # )
        logger.error(f"Message ID {message_id} not found for deletion.")
        return JsonResponse({'error': 'Message not found'}, status=404)


def extract_keywords(question):
    """Extracts keywords from a given question."""
    logger.info("Extracting keywords from question.")
    keywords = re.findall(r'\b\w+-\w+\b', question)
    question_cleaned = re.sub(r'\b\w+-\w+\b', '', question)
    doc = nlp(question_cleaned)

    for token in doc:
        if token.text.lower() not in custom_stop_words:
            if token.pos_ == "VERB":
                keywords.append(token.text)
            elif token.pos_ == "ADJ" and not token.is_stop:
                adj = morph.parse(token.text)[0]
                adj_feminine = adj.inflect({"femn", "sing", "nomn"})
                if adj_feminine:
                    keywords.append(adj_feminine.word)
                else:
                    keywords.append(token.lemma_)
            elif token.is_alpha and not token.is_stop and token.text.lower() not in custom_stop_words:
                keywords.append(token.lemma_)

    unique_keywords = list(set(keywords))
    logger.debug(f"Extracted keywords: {unique_keywords}")
    return unique_keywords


@csrf_exempt
def extract_keywords_view(request):
    """API endpoint to extract keywords from a question."""
    if request.method == "POST":
        logger.info("Extracting keywords via POST request.")
        try:
            data = json.loads(request.body)
            question = data.get('question', '')

            if not question:
                logger.warning("No question provided.")
                return JsonResponse({'error': 'Question not provided'}, status=400)

            keywords = extract_keywords(question)
            logger.info(f"Keywords extracted: {keywords}")
            return JsonResponse({'keywords': keywords})
        except json.JSONDecodeError:
            logger.error("Invalid JSON format received.")
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    logger.warning("Invalid method: only POST is supported.")
    return JsonResponse({'error': 'Invalid request method'}, status=405)

@csrf_exempt
def delete_relation(request):
    """Deletes a relation (edge) between two nodes in the OrientDB database."""

    if request.method != 'POST':
        logger.warning(f"Invalid request method: {request.method}")
        return JsonResponse({'error': 'Only POST method is allowed'}, status=405)

    try:
        data = json.loads(request.body)
        answer = data.get('answer_id')
        document = data.get('document_id')

        if not answer or not document:
            logger.warning("Missing 'answer' or 'document' fields.")
            return JsonResponse({'error': 'Missing required fields'}, status=400)

        # Optional: Validate RID format like #12:0 (OrientDB record ID)
        rid_pattern = r'^#\d+:\d+$'
        if not re.match(rid_pattern, answer) or not re.match(rid_pattern, document):
            logger.warning(f"Invalid RID format: answer={answer}, document={document}")
            return JsonResponse({'error': 'Invalid RID format'}, status=400)

        command = f"DELETE EDGE Includes WHERE out = {answer} AND in = {document}"
        headers = {'Content-Type': 'application/json'}
        json_data = {"command": command}

        response = requests.post(
            config_settings.ORIENT_COMMAND_URL,
            headers=headers,
            json=json_data,
            auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS),
            proxies={"http": None, "https": None}
        )

        if response.status_code != 200:
            logger.error(f"Failed to delete relation: {response.status_code}, {response.text}")
            return JsonResponse({'error': 'Failed to delete relation'}, status=500)

        logger.info(f"Relation deleted between {answer} and {document}")
        return JsonResponse({'message': 'Relation successfully deleted'}, status=200)

    except json.JSONDecodeError:
        logger.exception("Invalid JSON in request body")
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    except Exception as e:
        logger.exception("An unexpected error occurred while deleting the relation.")
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@csrf_exempt
def get_nodes_by_type(request):
    if request.method != 'GET':
        logger.warning(f"Invalid request method: {request.method}")
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    try:
        node_type = request.GET.get('type', '').strip()
        if not node_type:
            logger.warning("Node type parameter missing")
            return JsonResponse({'error': 'Node type parameter is required'}, status=400)

        node_type = urllib.parse.unquote(node_type)

        # Enhanced validation with allowlist approach
        valid_classes = ['document', 'link', 'Topic', 'Question', 'Section']  # Add all valid class names
        if node_type not in valid_classes:
            logger.warning(f"Invalid node type requested: {node_type}")
            return JsonResponse({'error': 'Invalid node type'}, status=400)

        # Additional length validation
        if len(node_type) > 50:
            logger.warning(f"Node type parameter too long: {node_type[:50]}...")
            return JsonResponse({'error': 'Node type parameter too long'}, status=400)

        # Parameterized query using OrientDB's parameter binding
        query = "SELECT FROM :node_type LIMIT -1"

        response = requests.post(
            config_settings.ORIENT_COMMAND_URL,
            auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS),
            headers={'Content-Type': 'application/json'},
            json={
                "command": query,
                "parameters": {"node_type": node_type}
            },
            timeout=10,  # Add timeout
            proxies={"http": None, "https": None}
        )

        if response.status_code != 200:
            logger.error(f"Database error ({response.status_code}) for node_type={node_type}")
            return JsonResponse({
                'error': 'Error fetching data from database',
                'status_code': response.status_code
            }, status=500)

        try:
            data = response.json()
        except ValueError as e:
            logger.error(f"Error parsing JSON response: {e}")
            return JsonResponse({'error': 'Failed to parse database response'}, status=500)

        results = data.get('result')
        if not isinstance(results, list):
            logger.error("Invalid response format: 'result' field missing or not a list")
            return JsonResponse({'error': 'Invalid database response format'}, status=500)

        # Sanitize output - remove internal fields
        sanitized_results = []
        for item in results:
            if not isinstance(item, dict):
                continue

            # Create a safe copy of the item
            sanitized = {k: v for k, v in item.items() if not k.startswith('@')}
            if '@rid' in item:
                sanitized['id'] = item['@rid']
            sanitized_results.append(sanitized)

        return JsonResponse({
            'status': 'success',
            'count': len(sanitized_results),
            'result': sanitized_results
        }, status=200)

    except requests.Timeout:
        logger.error("Database connection timed out")
        return JsonResponse({'error': 'Database connection timeout'}, status=504)
    except requests.RequestException as e:
        logger.error(f"Database connection error: {str(e)}")
        return JsonResponse({'error': 'Database connection error'}, status=500)
    except Exception as e:
        logger.exception(f"Unexpected error in get_nodes_by_type: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


@csrf_exempt
def create_relation(request):
    if request.method != 'POST':
        logger.warning("Invalid request method for relation creation")
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    try:
        data = json.loads(request.body)
        start_node_id = data.get('start_node_id')
        end_node_id = data.get('end_node_id')

        # Validate required fields
        if not start_node_id or not end_node_id:
            logger.warning("Missing required fields for relation creation")
            return JsonResponse({'error': 'Missing required fields: start_node_id and end_node_id'}, status=400)

        # Check if relation already exists using parameterized query
        check_response = requests.post(
            config_settings.ORIENT_COMMAND_URL,
            headers={'Content-Type': 'application/json'},
            json={
                "command": "SELECT FROM Includes WHERE out = :start_id AND in = :end_id",
                "parameters": {
                    "start_id": start_node_id,
                    "end_id": end_node_id
                }
            },
            auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS)
        )

        # Handle check response
        if not check_response.ok:
            logger.error(f"Failed to check relation existence: {check_response.text}")
            return JsonResponse({'error': 'Failed to verify relation status'}, status=500)

        check_data = check_response.json()
        if check_data.get('result') and len(check_data['result']) > 0:
            logger.warning(f"Relation exists between {start_node_id} and {end_node_id}")
            return JsonResponse({
                'error': 'Relation already exists',
                'existing_relation': check_data['result'][0]
            }, status=409)

        # Create the relation using parameterized query
        create_response = requests.post(
            config_settings.ORIENT_COMMAND_URL,
            headers={'Content-Type': 'application/json'},
            json={
                "command": "CREATE EDGE Includes FROM :start_id TO :end_id",
                "parameters": {
                    "start_id": start_node_id,
                    "end_id": end_node_id
                }
            },
            auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS)
        )

        # Handle create response
        if create_response.ok:
            result = create_response.json().get('result', {})
            logger.info(f"Created relation between {start_node_id} and {end_node_id}")
            return JsonResponse({
                'message': 'Relation successfully created',
                'relation': result
            }, status=201)
        else:
            logger.error(f"Failed to create relation: {create_response.text}")
            return JsonResponse({
                'error': 'Failed to create relation',
                'details': create_response.text
            }, status=500)

    except json.JSONDecodeError:
        logger.error("Invalid JSON format in request body")
        return JsonResponse({'error': 'Invalid request format'}, status=400)
    except requests.RequestException as e:
        logger.error(f"Database connection error: {str(e)}")
        return JsonResponse({'error': 'Database operation failed'}, status=500)
    except Exception as e:
        logger.exception(f"Unexpected error in relation creation: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


@csrf_exempt
def delete_node(request):
    """Creates a relation between two nodes."""
    user = request.user
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            node_id_to_delete = data.get('node_id')

            if not node_id_to_delete:
                logger.warning("Missing required fields for relation creation.")
                return JsonResponse({'error': 'Missing required fields'}, status=400)

            command = f"DELETE VERTEX {node_id_to_delete}"
            headers = {'Content-Type': 'application/json'}
            json_data = {"command": command}

            response = requests.post(config_settings.ORIENT_COMMAND_URL, headers=headers, json=json_data,
                                     auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS))

            log_action(f"Удаление вершины {node_id_to_delete}.")
            return JsonResponse({'message': 'Node successfully deleted'}, status=201)
        except Exception as e:

            logger.exception("An error occurred while creating a relation.")
            return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def get_nodes(request):
    """Retrieves all nodes."""
    if request.method == 'GET':
        logger.info("Fetching all nodes.")

        try:
            sql_command = f"SELECT * FROM V"
            headers = {'Content-Type': 'application/json'}
            json_data = {"command": sql_command}

            response = requests.post(config_settings.ORIENT_COMMAND_URL, headers=headers, json=json_data,
                                     auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS))

            if response.status_code == 200:
                logger.info(f"Nodes get successfully: {response.text}")
                try:
                    response_data = response.json()
                    logger.info(f"Response JSON: {response_data}")
                    return JsonResponse({'status': 'success', 'data': response_data}, status=201)
                except ValueError as e:
                    logger.error(f"Error parsing JSON response: {e}")
                    return JsonResponse({'error': 'Failed to parse response'}, status=500)

            else:
                logger.error(f"Error fetching data: HTTP {response.status_code} - {response.text}")
                return JsonResponse({'error': f"Error {response.status_code}: {response.text}"},
                                    status=response.status_code)
        except Exception as e:
            logger.exception("An error occurred while fetching nodes.")
            return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def create_training_message(request):
    """Creates a new training message."""
    logger.info("Creating a new training message.")
    user = request.user
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            sender_id = data.get('sender_id')
            content = data.get('content')
            neural_message = data.get('neural_message')
            recognized_message = data.get('recognized_message')

            if not content:
                logger.warning("Content field is missing.")
                return JsonResponse({'error': 'Field "content" is required.'}, status=400)

            sender = None
            if sender_id:
                logger.debug(f"Fetching user with ID: {sender_id}")
                try:
                    sender = ChatUser.objects.get(id=sender_id)
                except ChatUser.DoesNotExist:
                    logger.error(f"User with ID {sender_id} not found.")
                    return JsonResponse({'error': 'User not found.'}, status=404)

            training_message = TrainingMessage.objects.create(
                sender=sender,
                neural_message = neural_message,
                recognized_message = recognized_message,
                content=content
            )

            logger.info(f"Training message created with ID: {training_message.id}")
            # user_action.info(
            #     f"Training message created with ID: {training_message.id}",
            #     extra={
            #         'user_id': user.id,
            #         'user_name': user.first_name + ' ' + user.last_name,
            #         'action_type': 'create_training_message',
            #         'time': datetime.now(),
            #         'details': json.dumps({
            #             'status': f"Training message created with ID: {training_message.id}",
            #         })
            #
            #     }
            # )

            return JsonResponse({
                'id': training_message.id,
                'sender': sender_id,
                'content': training_message.content,
                'neural_message': neural_message if neural_message else '',
                'recognized_message': recognized_message if recognized_message else '',
                'created_at': training_message.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'is_ignored': training_message.is_ignored,
                'is_unread': training_message.is_unread
            }, status=201)

        except json.JSONDecodeError:
            logger.error("Invalid JSON format received.")
            return JsonResponse({'error': 'Invalid JSON format.'}, status=400)
        except Exception as e:
            logger.exception("An error occurred while creating a training message.")
            return JsonResponse({'error': str(e)}, status=500)

    logger.warning("Invalid method: only POST is supported.")
    return JsonResponse({'error': 'Method not supported. Use POST.'}, status=405)

# chat_dashboard/views.py

#@role_required(['admin',])
def user_list(request):
    """Displays a combined list of User and ChatUser."""
    logger.info("Accessing user list.")

    search_query = request.GET.get('search', '')
    sort_column = request.GET.get('sort', 'username')
    page_number = request.GET.get('page', 1)
    archive_filter = request.GET.get('archive_filter', 'all')

    print(archive_filter)

    # Получаем пользователей с учетом фильтра архива
    users = User.objects.all()
    chat_users = ChatUser.objects.all()

    if archive_filter == 'active':
        users = users.filter(is_archived=False)
        chat_users = chat_users.filter(is_archived=False)
    elif archive_filter == 'archived':
        users = users.filter(is_archived=True)
        chat_users = chat_users.filter(is_archived=True)
    # Для 'all' - оставляем всех пользователей без фильтрации

    combined_users = []
    for user in users:
        combined_users.append({
            'type': 'admin',
            'id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'username': user.username,
            'email': user.email,
            'role': user.get_role_display(),
            'is_archived': user.is_archived
        })

    for chat_user in chat_users:
        combined_users.append({
            'type': 'chat',
            'id': chat_user.id,
            'first_name': chat_user.first_name,
            'last_name': chat_user.last_name,
            'username': chat_user.username,
            'email': chat_user.email,
            'role': chat_user.get_role_display(),
            'is_archived': chat_user.is_archived
        })

    # Фильтрация по поисковому запросу
    if search_query:
        combined_users = [
            u for u in combined_users
            if (search_query.lower() in u['first_name'].lower() or
                search_query.lower() in u['last_name'].lower() or
                search_query.lower() in u['email'].lower() or
                search_query.lower() in u['username'].lower())
        ]

    # Сортировка
    reverse_sort = sort_column.startswith('-')
    sort_key = sort_column.lstrip('-') if reverse_sort else sort_column

    combined_users.sort(
        key=lambda x: str(x.get(sort_key, '')).lower(),
        reverse=reverse_sort
    )

    paginator = Paginator(combined_users, 10)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    return render(request, 'chat_dashboard/users.html', {
        'page_obj': page_obj,
        'sort_column': sort_column,
        'search_query': search_query,
        'archive_filter': archive_filter  # Передаём текущий фильтр в шаблон
    })

@csrf_exempt
#@role_required('admin')
def archive_user(request, user_type, user_id):
    """Archive a user instead of deleting."""
    if user_type == 'admin':
        user = get_object_or_404(User, id=user_id)
    else:
        user = get_object_or_404(ChatUser, id=user_id)

    if request.method == 'POST':
        user.is_archived = True
        user.save()
        messages.success(request, f'Пользователь {user.username} перемещен в архив.')
        return redirect('chat_dashboard:user_list')

    return redirect('chat_dashboard:user_list')

@csrf_exempt
#@role_required('admin')
def restore_user(request, user_type, user_id):
    """Restore an archived user."""
    if user_type == 'admin':
        user = get_object_or_404(User, id=user_id)
    else:
        user = get_object_or_404(ChatUser, id=user_id)

    if request.method == 'POST':
        user.is_archived = False
        user.save()
        messages.success(request, f'Пользователь {user.username} восстановлен из архива.')
        return redirect('chat_dashboard:user_list')

    return redirect('chat_dashboard:user_list')

#@role_required('admin')
@csrf_exempt
def user_create(request):
    """Creates a new user."""
    logger.info("Creating a new user.")
    form = UserForm(request.POST or None)
    user = request.user
    if request.method == 'POST':
        email = request.POST.get('email')
        if User.objects.filter(email=email).exists():
            logger.error(f"Attempt to create user failed: Email already registered - {email}")
            # Возвращаем информацию о том, что email уже зарегистрирован
            log_action(f"Ошибка создания нового пользователя. Почта {email} уже зарегистрирована.")
            return render(request, 'chat_dashboard/user_create_form.html', {
                'form': form,
                'email_exists': True,
                'email': email
            })

        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            #user.is_active = True
            user.save()
            logger.info(f"User created: ID={user.id}, Username={user.username}, Email={user.email}")
            log_action(f"Создание нового пользователя. ID: {user.id}, email: {user.email}, role: {user.get_role_display()}.")
            # user_action.info(
            #     f"User created: ID={user.id}, Username={user.username}, Email={user.email}",
            #     extra={
            #         'user_id': user.id,
            #         'user_name': user.first_name + ' ' + user.last_name,
            #         'action_type': 'create_user',
            #         'time': datetime.now(),
            #         'details': json.dumps({
            #             'status': f"User created: ID={user.id}, Username={user.username}, Email={user.email}",
            #         })
            #     })
            messages.success(request,
                             "Создана новая учетная запись. Данные для её активации направлены на указанный вами электронный адрес.")
            return redirect('chat_dashboard:user_list')  # Перенаправляем на список пользователей

    return render(request, 'chat_dashboard/user_create_form.html', {'form': form})

def get_user_model_by_type(user_type):
    if user_type == 'admin':
        return get_user_model()
    elif user_type == 'chat':
        return ChatUser
    raise Http404("Invalid user type")


#@role_required('admin')
@csrf_exempt
def user_update(request, user_type, pk):
    """Updates user data for both User and ChatUser models."""
    logger.info(f"Updating {user_type} user with ID: {pk}")
    model = get_user_model_by_type(user_type)
    user = get_object_or_404(model, pk=pk)

    if request.method == 'POST':
        form = UserFormUpdate(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)
            if user_type == 'admin' and form.cleaned_data.get('password'):
                user.set_password(form.cleaned_data['password'])
            user.save()
            logger.info(f"{user_type.capitalize()} user updated: ID={pk}")
            log_action(f"Обновление данных пользователя. ID: {pk}, email: {user.email}, role: {user.get_role_display()}")
            # user_action.info(
            #     f"{user_type.capitalize()} user updated: ID={pk}",
            #     extra={
            #         'user_id': user.id,
            #         'user_name': user.first_name + ' ' + user.last_name,
            #         'action_type': 'update_user',
            #         'time': datetime.now(),
            #         'details': json.dumps({
            #             'status': f"{user_type.capitalize()} user updated: ID={pk}",
            #         })
            #     })
            return redirect('chat_dashboard:user_list')
    else:
        form = UserFormUpdate(instance=user)

    return render(request, 'chat_dashboard/user_update_form.html', {
        'form': form,
        'user_type': user_type
    })

#@role_required('admin')
@csrf_exempt
def user_delete(request, user_type, pk):
    """Deletes a user from specified model."""
    logger.info(f"Attempting to delete {user_type} user with ID: {pk}")
    model = get_user_model_by_type(user_type)
    user = get_object_or_404(model, pk=pk)

    if request.method == 'POST':
        user.delete()
        logger.info(f"user deleted: ID={pk}")
        log_action(f"Удаление пользователя. ID: {pk}.")
        return redirect('chat_dashboard:user_list')

    return render(request, 'chat_dashboard/user_delete_form.html', {
        'user': user,
        'user_type': user_type
    })


def get_last_message_subquery(field):
    logger.debug("Creating a subquery for the last message.")
    return Message.objects.filter(dialog=OuterRef('pk')).order_by('-created_at').values(field)[:1]


#@role_required(['admin', 'operator'])
def archive(request):
    user = request.user
    logger.info(f"Accessing archive page by user {user}.")
    if request.method == 'POST':
        user.delete()

    dialogs = Dialog.objects.annotate(
        has_messages=Exists(Message.objects.filter(dialog=OuterRef('pk'))),
        username=Concat(
            F('user__first_name'),
            Value(' '),
            F('user__last_name')
        ),
        last_message=Subquery(
            Message.objects.filter(dialog=OuterRef('pk')).order_by('-created_at').values('content')[:1]
        ),
        last_message_timestamp=Subquery(
            Message.objects.filter(dialog=OuterRef('pk')).order_by('-created_at').values('created_at')[:1]
        ),
        last_message_sender_id=Subquery(
            Message.objects.filter(dialog=OuterRef('pk')).order_by('-created_at').values('sender_id')[:1]
        ),
        last_message_username=Case(
            When(last_message_sender_id=None, then=Value('Bot')),
            default=Subquery(
                Message.objects.filter(dialog=OuterRef('pk')).order_by('-created_at').values('sender__first_name')[:1]
            )
        ),
    ).filter(has_messages=True).order_by('-last_message_timestamp')

    # Логирование для отладки
    logger.debug(f"Dialogs retrieved: {list(dialogs)}")

    return render(request, 'chat_dashboard/archive.html', {
        'dialogs': dialogs,
        'user': user,
    })


#@role_required(['admin', 'operator'])
def create_or_edit_content(request):
    user = request.user
    # user_action.info(
    #     f"Accessing create_or_edit_content page by user {user}.",
    #     extra={
    #         'user_id': user.id,
    #         'user_name': user.first_name + ' ' + user.last_name,
    #         'action_type': 'access to create_or_edit_content',
    #         'time': datetime.now(),
    #         'details': json.dumps({
    #             'status': f"Accessing create_or_edit_content page by user {user}.",
    #         })
    #     })
    return render(request, 'chat_dashboard/edit_content.html')


def filter_dialogs_by_date_range(request):
    user = request.user
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')

    # Преобразование в aware datetime
    start = timezone.make_aware(timezone.datetime.fromisoformat(start_date))
    end = timezone.make_aware(timezone.datetime.fromisoformat(end_date))

    dialogs = Dialog.objects.annotate(
        has_messages=Exists(Message.objects.filter(dialog=OuterRef('pk'))),
        username=Concat(F('user__first_name'), Value(' '), F('user__last_name')),
        last_message=Subquery(
            Message.objects.filter(dialog=OuterRef('pk')).order_by('-created_at').values('content')[:1]),
        last_message_timestamp=Subquery(
            Message.objects.filter(dialog=OuterRef('pk')).order_by('-created_at').values('created_at')[:1]),
        last_message_sender_id=Subquery(
            Message.objects.filter(dialog=OuterRef('pk')).order_by('-created_at').values('sender_id')[:1]),
        last_message_username=Case(
            When(last_message_sender_id=None, then=Value('Bot')),
            default=Subquery(
                Message.objects.filter(dialog=OuterRef('pk')).order_by('-created_at').values('sender__first_name')[:1])
        ),
    ).filter(has_messages=True, last_message_timestamp__range=(start, end)).order_by('-last_message_timestamp')

    logger.debug(f"Filtered dialogs: {list(dialogs)}")

    # Возвращаем отфильтрованные диалоги в формате JSON
    dialogs_data = [
        {
            'id': dialog.id,
            'user': {
                'id': dialog.user.id,
                'username': dialog.user.username
            },
            'last_message': dialog.last_message,
            'last_message_timestamp': dialog.last_message_timestamp,
            'last_message_username': dialog.last_message_username
        }
        for dialog in dialogs
    ]

    return JsonResponse(dialogs_data, safe=False)


def filter_dialogs(request):
    period = request.GET.get('period', 0)
    user_id = request.GET.get('user_id')
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    last_name = request.GET.get('last_name')
    content = request.GET.get('content')
    rating_type = request.GET.get('rating_type')

    try:
        period = int(period)
    except ValueError:
        period = 0

    dialogs = Dialog.objects.annotate(
        has_messages=Exists(Message.objects.filter(dialog=OuterRef('pk'))),
        last_message=Subquery(
            Message.objects.filter(
                dialog=OuterRef('pk'),
                message_type__in=['message', 'document', 'link', 'operator']  # Exclude like/dislike from last message
            )
            .order_by('-created_at')
            .values('content')[:1]
        ),
        last_message_timestamp=Subquery(
            Message.objects.filter(dialog=OuterRef('pk'))
            .order_by('-created_at')
            .values('created_at')[:1]
        ),
        last_message_sender_id=Subquery(
            Message.objects.filter(
                dialog=OuterRef('pk'),
                message_type__in=['message', 'document', 'link',]  # Exclude like/dislike from sender
            )
            .order_by('-created_at')
            .values('sender_id')[:1]
        ),
        last_message_username=Case(
            When(last_message_sender_id__isnull=True, then=Value('Bot')),
            default=Concat(
                Subquery(
                    ChatUser.objects.filter(pk=OuterRef('last_message_sender_id'))
                    .values('first_name')[:1]
                ),
                Value(' '),
                Subquery(
                    ChatUser.objects.filter(pk=OuterRef('last_message_sender_id'))
                    .values('last_name')[:1]
                ),
            ),
            output_field=CharField()
        ),
        username=Concat(
            F('user__first_name'),
            Value(' '),
            F('user__last_name'),
            output_field=CharField()
        ),
        # Updated annotations for ratings using message_type
        has_likes=Exists(
            Message.objects.filter(
                dialog=OuterRef('pk'),
                message_type='like'
            )
        ),
        has_dislikes=Exists(
            Message.objects.filter(
                dialog=OuterRef('pk'),
                message_type='dislike'
            )
        ),
        has_any_rating=Exists(
            Message.objects.filter(
                dialog=OuterRef('pk'),
                message_type__in=['like', 'dislike']
            )
        )
    ).filter(has_messages=True)

    if user_id:
        dialogs = dialogs.filter(user_id=user_id)

    if last_name:
        dialogs = dialogs.filter(user__last_name__icontains=last_name)

    if period > 0:
        time_filter = timezone.now() - timedelta(days=period)
        dialogs = dialogs.filter(last_message_timestamp__gte=time_filter)

    if start_date and end_date:
        try:
            start = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1)
            dialogs = dialogs.filter(last_message_timestamp__range=(start, end))
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)

    if content:
        dialogs = dialogs.filter(
            Exists(
                Message.objects.filter(
                    dialog=OuterRef('pk'),
                    content__icontains=content,
                    message_type__in=['message', 'document', 'link', 'operator']  # Only search in actual messages
                )
            )
        )

    if rating_type:
        if rating_type == 'like':
            dialogs = dialogs.filter(has_likes=True)
        elif rating_type == 'dislike':
            dialogs = dialogs.filter(has_dislikes=True)
        elif rating_type == 'any':
            dialogs = dialogs.filter(has_any_rating=True)
        elif rating_type == 'none':
            dialogs = dialogs.filter(has_any_rating=False)

    dialogs = dialogs.order_by('-last_message_timestamp').distinct()

    dialogs_data = [
        {
            'id': dialog.id,
            'user': {
                'id': dialog.user.id,
                'username': dialog.username,
                'last_name': dialog.user.last_name,
                'first_name': dialog.user.first_name,
                'tabel_number': dialog.user.tabel_number
            },
            'last_message': dialog.last_message,
            'last_message_timestamp': dialog.last_message_timestamp,
            'last_message_username': dialog.last_message_username,
            'has_likes': dialog.has_likes,
            'has_dislikes': dialog.has_dislikes
        }
        for dialog in dialogs
    ]

    return JsonResponse(dialogs_data, safe=False)


def export_to_excel(request):
    # Получаем параметры фильтрации
    period = request.GET.get('period', 0)
    user_id = request.GET.get('user_id')
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    last_name = request.GET.get('last_name')
    content = request.GET.get('content')
    rating_type = request.GET.get('rating_type')

    try:
        period = int(period)
    except ValueError:
        period = 0

    # Базовый запрос с аннотациями
    dialogs = Dialog.objects.annotate(
        has_messages=Exists(Message.objects.filter(dialog=OuterRef('pk'))),
        last_message=Subquery(
            Message.objects.filter(
                dialog=OuterRef('pk'),
                message_type__in=['message', 'document', 'link', 'operator']
            )
            .order_by('-created_at')
            .values('content')[:1]
        ),
        last_message_timestamp=Subquery(
            Message.objects.filter(dialog=OuterRef('pk'))
            .order_by('-created_at')
            .values('created_at')[:1]
        ),
        # Упрощаем получение имени отправителя
        last_message_sender=Case(
            When(messages__message_type__in=['message', 'document', 'link', 'operator'],
                 then=Concat(
                     F('messages__sender__first_name'),
                     Value(' '),
                     F('messages__sender__last_name')
                 )),
            default=Value('Bot'),
            output_field=CharField()
        ),
        has_likes=Exists(
            Message.objects.filter(
                dialog=OuterRef('pk'),
                message_type='like'
            )
        ),
        has_dislikes=Exists(
            Message.objects.filter(
                dialog=OuterRef('pk'),
                message_type='dislike'
            )
        )
    ).filter(has_messages=True)

    # Применяем фильтры
    if user_id:
        dialogs = dialogs.filter(user_id=user_id)
    if last_name:
        dialogs = dialogs.filter(user__last_name__icontains=last_name)
    if period > 0:
        time_filter = timezone.now() - timedelta(days=period)
        dialogs = dialogs.filter(last_message_timestamp__gte=time_filter)
    if start_date and end_date:
        try:
            start = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d')) + timedelta(days=1)
            dialogs = dialogs.filter(last_message_timestamp__range=(start, end))
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)
    if content:
        dialogs = dialogs.filter(
            Exists(
                Message.objects.filter(
                    dialog=OuterRef('pk'),
                    content__icontains=content,
                    message_type__in=['message', 'document', 'link', 'operator']
                )
            )
        )
    if rating_type:
        if rating_type == 'like':
            dialogs = dialogs.filter(has_likes=True)
        elif rating_type == 'dislike':
            dialogs = dialogs.filter(has_dislikes=True)
        elif rating_type == 'any':
            dialogs = dialogs.filter(has_likes=True) | dialogs.filter(has_dislikes=True)
        elif rating_type == 'none':
            dialogs = dialogs.exclude(has_likes=True).exclude(has_dislikes=True)

    # Получаем данные с нужными полями
    dialogs = dialogs.select_related('user').order_by('-last_message_timestamp').distinct()

    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "История чатов"

    # Заголовки
    headers = [
        "ID диалога",
        "Фамилия",
        "Имя",
        "Табельный номер",
        "Дата последнего сообщения",
        "Последнее сообщение",
        "Отправитель последнего сообщения",
        "Есть лайки",
        "Есть дизлайки"
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Заполняем данные
    for row_num, dialog in enumerate(dialogs, 2):
        ws[f"A{row_num}"] = dialog.id
        ws[f"B{row_num}"] = dialog.user.last_name
        ws[f"C{row_num}"] = dialog.user.first_name
        ws[f"D{row_num}"] = dialog.user.tabel_number
        ws[f"E{row_num}"] = dialog.last_message_timestamp.strftime(
            '%Y-%m-%d %H:%M:%S') if dialog.last_message_timestamp else ''
        ws[f"F{row_num}"] = dialog.last_message if dialog.last_message else ''
        ws[f"G{row_num}"] = dialog.last_message_sender if hasattr(dialog, 'last_message_sender') else 'Bot'
        ws[f"H{row_num}"] = "Да" if dialog.has_likes else "Нет"
        ws[f"I{row_num}"] = "Да" if dialog.has_dislikes else "Нет"

    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Формируем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="chat_history.xlsx"'},
    )
    wb.save(response)

    return response


def get_messages(request, dialog_id):
    """Retrieves all messages in a dialog."""
    logger.info(f"Fetching messages for dialog ID: {dialog_id}")

    try:
        messages = Message.objects.filter(dialog_id=dialog_id).order_by('created_at')

        # Если сообщений нет
        if not messages:
            logger.error(f"No messages found for dialog ID {dialog_id}.")
            return JsonResponse({"status": "error", "message": "No messages found."}, status=404)

        logger.debug(f"Messages retrieved: {messages}")
        messages_data = [
            {
                'sender': (
                    f"{message.sender.first_name} {message.sender.last_name}".strip()
                    if message.sender and message.sender_type == 'user'
                    else 'bot'
                ),
                'message_type': message.message_type,
                'content': message.content,
                'timestamp': message.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
            for message in messages
        ]
        logger.debug(f"Messages retrieved: {messages_data}")
        return JsonResponse({'messages': messages_data})

    except Exception as e:
        logger.error(f"Error fetching messages for dialog ID {dialog_id}: {str(e)}")
        return JsonResponse({"status": "error", "message": "Internal server error."}, status=500)


def get_info(request, user_id):
    try:
        user = ChatUser.objects.get(id=user_id)
        active_session = Session.objects.filter(user=user, expires_at__gt=now()).first()
        if active_session:
            user_status = {
                'is_online': True,
                'last_active': 'Недавно'
            }
        else:
            last_session = Session.objects.filter(user=user).order_by('-expires_at').first()
            user_status = {
                'is_online': False,
                'last_active': last_session.expires_at if last_session else 'Неизвестно'
            }
    except ChatUser.DoesNotExist:
        user_status = {
            'is_online': False,
            'last_active': 'Неизвестно'
        }

    return JsonResponse({'status': user_status})


#@role_required(['admin'])
def settings_view(request):
    settings, created = Settings.objects.get_or_create(id=1)
    user = request.user
    # user_action.info(
    #     f"Accessing settings page by user {user}.",
    #     extra={
    #         'user_id': user.id,
    #         'user_name': user.first_name + ' ' + user.last_name,
    #         'action_type': 'access to settings page',
    #         'time': datetime.now(),
    #         'details': json.dumps({
    #             'status': f"Accessing settings page by user {user}.",
    #         })
    #     })
    months = list(range(1, 25))
    current_retention_months = settings.message_retention_days // 30 if settings.message_retention_days else 1

    return render(request, 'chat_dashboard/settings.html', {
        'settings': settings,
        'months': months,
        'current_retention_months': current_retention_months,
        'ip_address': settings.ip_address,
        'logs_backup': settings.logs_backup
    })

@csrf_exempt
def update_settings(request):
    if request.method == 'POST':
        try:
            session_duration = int(request.POST.get('session_duration'))
            retention_months = request.POST.get('message_retention_months')
            logs_backup = request.POST.get('logs_backup')
            neural_active = request.POST.get('neural_active') == 'on'
            settings = Settings.objects.first()
            settings.session_duration = session_duration
            settings.logs_backup = logs_backup
            settings.neural_active = neural_active
            settings.message_retention_days = int(retention_months) * 30 if retention_months.isdigit() else settings.message_retention_days
            settings.save()

            return JsonResponse({'status': 'success', 'message': 'Настройки успешно обновлены!'})

        except Exception as e:

            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Неверный запрос'})


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


@csrf_exempt
def upload_document(request):
    user = request.user
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name
        file_path = os.path.join(settings.MEDIA_ROOT, 'documents', file_name)
        document_uuid = str(uuid.uuid4())

        # Проверка, существует ли файл локально
        if os.path.exists(file_path):
            logger.warning(f"Document '{file_name}' already exists locally.")
            return JsonResponse(
                {'message': 'Документ уже существует локально!', 'data': {'file_name': file_name}},
                status=400
            )

        # Проверка, существует ли документ в базе данных
        if Documents.objects.filter(document_name=file_name).exists():
            logger.warning(f"Document '{file_name}' already exists in the database.")
            return JsonResponse(
                {'message': 'Документ уже существует в базе данных!', 'data': {'file_name': file_name}},
                status=400
            )

        # Если документ не существует, создаем его
        try:
            new_document = Documents.objects.create(
                document_name=file_name,
                document_uuid=document_uuid
            )

            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            logger.info(f"Document '{file_name}' successfully uploaded.")
            # user_action.info(
            #     f"Document '{file_name}' uploaded by user {user}.",
            #     extra={
            #         'user_id': user.id,
            #         'user_name': user.first_name + ' ' + user.last_name,
            #         'action_type': 'upload document',
            #         'time': datetime.now(),
            #         'details': json.dumps({
            #             'status': f"Document '{file_name}' uploaded by user {user}.",
            #         })
            #     }
            # )

            return JsonResponse(
                {'message': 'Файл успешно загружен!', 'data': {'file_name': file_name, 'file_id': document_uuid}},
                status=200
            )

        except Exception as e:
            logger.error(f"Error uploading document '{file_name}': {e}")
            return JsonResponse({'message': 'Ошибка при загрузке файла!'}, status=500)

    # Если файл не был передан
    logger.warning("No file provided for upload.")
    return JsonResponse({'message': 'Файл не загружен!'}, status=400)

    logger.warning("No file provided for upload.")
    return JsonResponse({'message': 'Файл не загружен!'}, status=400)


def get_document_link_by_uuid(request, uuid):
    documents_path = os.path.join(settings.MEDIA_ROOT, 'documents')
    available_files = os.listdir(documents_path)

    document = Documents.objects.get(document_uuid=uuid)
    document_name = document.document_name

    for file in available_files:
        if document_name.strip() == file.strip():
            file_url = f"{settings.MEDIA_URL}documents/{file}"

            return JsonResponse({'file_url': file_url}, status=200)

    return JsonResponse({'error': 'Файл не найден'}, status=404)

def get_documents(request):
    if request.method == 'GET':
        logger.info("Fetching all nodes of type 'document'.")
        url = f"{config_settings.ORIENT_QUERY_URL}/SELECT * FROM document LIMIT -1"

        try:
            response = requests.get(
                url,
                auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS),
                headers={"Accept": "application/json"}
            )
            if not response.ok:
                logger.warning("Failed to fetch data from OrientDB.")
                return JsonResponse([], safe=False, status=200)

            data = response.json()
            if 'result' in data:
                documents_data = []
                for document in data['result']:
                    if 'name' in document and '@rid' in document:
                        documents_data.append({
                            'id': document['@rid'],
                            'name': document['name']
                        })

                logger.info(f"Found {len(documents_data)} documents.")
                return JsonResponse({'result': documents_data}, safe=False)
            else:
                logger.error("Unexpected response format.")
                return JsonResponse({"error": "Unexpected response format"}, status=500)

        except Exception as e:
            logger.error(f"Error fetching data: {e}")
            return JsonResponse({"error": "Failed to fetch data"}, status=500)


def get_links(request):
    if request.method == 'GET':
        logger.info("Fetching all nodes of type 'link'.")
        url = f"{config_settings.ORIENT_QUERY_URL}/SELECT * FROM link LIMIT -1"

        try:
            response = requests.get(
                url,
                auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS),
                headers={"Accept": "application/json"}
            )
            if not response.ok:
                logger.warning("Failed to fetch data from OrientDB.")
                return JsonResponse([], safe=False, status=200)

            data = response.json()
            if 'result' in data:
                links_data = []
                for link in data['result']:
                    if 'name' in link and '@rid' in link:
                        links_data.append({
                            'id': link['@rid'],
                            'name': link['name']
                        })

                logger.info(f"Found {len(links_data)} documents.")
                return JsonResponse({'result': links_data}, safe=False)
            else:
                logger.error("Unexpected response format.")
                return JsonResponse({"error": "Unexpected response format"}, status=500)

        except Exception as e:
            logger.error(f"Error fetching data: {e}")
            return JsonResponse({"error": "Failed to fetch data"}, status=500)

@csrf_exempt
def add_question_to_existing(request):
    data = json.loads(request.body)
    existing_question = data.get('existing_question')
    print(existing_question)
    user_input = data.get('user_input')
    print(user_input)
    file_path = 'chat_user/questions.json'

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            questions_list = json.loads(f.read())

        found = False
        if user_input:
            for group in questions_list:
                for question in questions_list[group]:
                    if question == existing_question:
                        questions_list[group][question].append(user_input)
                        found = True
                        break
                if found:
                    break

        if not found:
            print(f"Вопрос '{existing_question}' не найден в группе.")
            return JsonResponse({"error": "Question not found"}, status=404)

        with open(file_path, 'w') as f:
            json.dump(questions_list, f, ensure_ascii=False, indent=4)
        print("Файл успешно обновлён!")
        return JsonResponse({"result": "Success"}, status=200)


    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)




@csrf_exempt
def add_new_question_from_teaching(request):
    data = json.loads(request.body)
    user_message = data.get('user_message')
    message_keywords = data.get('keywords')
    file_path = 'chat_user/questions.json'

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            questions_list = json.loads(f.read())

        questions_list[message_keywords] = {user_message: {}}

        with open(file_path, 'w') as f:
            json.dump(questions_list, f, ensure_ascii=False, indent=4)
        return JsonResponse({"result": "Success"}, status=200)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def delete_topic_relation(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=400)

    try:
        data = json.loads(request.body)
        topic = data.get('start_node_id')
        question = data.get('end_node_id')

        if not topic or not question:
            logger.warning("Missing required fields for relation deletion")
            return JsonResponse({'error': 'Missing required fields'}, status=400)

        # Параметризованный запрос
        command = "DELETE EDGE Includes WHERE out = :topic AND in = :question"
        params = {
            "topic": topic,
            "question": question
        }

        response = requests.post(
            config_settings.ORIENT_COMMAND_URL,
            headers={'Content-Type': 'application/json'},
            json={
                "command": command,
                "parameters": params
            },
            auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS),
            proxies={"http": None, "https": None}
        )

        # Проверка успешности операции
        if response.status_code == 200:
            result = response.json()
            if result.get('result'):
                deleted_count = len(result['result'])
                if deleted_count > 0:
                    logger.info(f"Successfully deleted {deleted_count} relation(s)")
                    return JsonResponse({'message': 'Relation successfully deleted'}, status=200)
                else:
                    logger.warning("No relations found to delete")
                    return JsonResponse({'error': 'Relation not found'}, status=404)

        logger.error(f"Database error: {response.text}")
        return JsonResponse({'error': 'Database operation failed'}, status=500)

    except json.JSONDecodeError:
        logger.error("Invalid JSON format in request")
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    except requests.RequestException as e:
        logger.error(f"Database connection error: {str(e)}")
        return JsonResponse({'error': 'Database connection error'}, status=500)
    except Exception as e:
        logger.exception("Unexpected error in delete_topic_relation")
        return JsonResponse({'error': 'Internal server error'}, status=500)

@csrf_exempt
def update_node(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            node_id = data.get('node_id')
            name = data.get('name')
            type = data.get('type')
            content = data.get('content', None)

            if not node_id or not name or not type:
                logger.error("Missing required fields")
                return JsonResponse({"error": "Missing required fields"}, status=400)

            # Подготовка параметров для безопасного запроса
            params = {
                'node_id': node_id,
                'name': name
            }

            if type == 'link':
                if content is None:
                    logger.error("Content is required for link nodes")
                    return JsonResponse({"error": "Content is required for link nodes"}, status=400)

                # Экранирование только для специальных нужд (не для безопасности!)
                escaped_content = content.replace("\u200b", "").replace("\n", "\\n").strip()
                params['content'] = escaped_content

                query = """
                    UPDATE link 
                    SET content = :content, name = :name 
                    WHERE @rid = :node_id
                """
            else:
                query = """
                    UPDATE document 
                    SET name = :name 
                    WHERE @rid = :node_id
                """

            # Отправляем параметризованный запрос
            response = requests.post(
                config_settings.ORIENT_COMMAND_URL,
                auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS),
                headers={"Content-Type": "application/json; charset=utf-8"},
                json={
                    "command": query,
                    "parameters": params
                },
            )

            logger.info(f"Response status: {response.status_code}, content: {response.text}")

            # Проверка успешности запроса
            if response.status_code != 200:
                logger.warning(f"Database error: {response.text}")
                return JsonResponse({"error": "Database operation failed"}, status=500)

            try:
                result = response.json()
                if result.get('result', []):
                    return JsonResponse({"message": "Successfully updated node"}, status=200)
                else:
                    logger.warning(f"Node with ID {node_id} not found")
                    return JsonResponse({"error": "Node not found"}, status=404)

            except ValueError:
                logger.error(f"Invalid JSON response: {response.text}")
                return JsonResponse({"error": "Invalid database response"}, status=500)

        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except requests.RequestException as e:
            logger.error(f"Request error: {e}")
            return JsonResponse({'error': 'Database connection error'}, status=500)
        except Exception as e:
            logger.error(f"Unexpected error: {e}")
            return JsonResponse({'error': 'Internal server error'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@csrf_exempt
def create_node(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    try:
        data = json.loads(request.body)
        node_class = data.get('class')
        node_name = data.get('name')
        node_content = data.get('content')
        node_uuid = data.get('uuid')

        # Валидация входных данных
        if not node_class or node_content is None:
            logger.warning("Missing required fields: class or content")
            return JsonResponse({'error': 'Missing required fields: class or content'}, status=400)

        # Проверка существующих узлов (только для класса 'link')
        if node_class == 'link':
            # Проверка по имени (если указано)
            if node_name:
                check_response = requests.post(
                    config_settings.ORIENT_COMMAND_URL,
                    headers={'Content-Type': 'application/json'},
                    json={
                        "command": "SELECT FROM link WHERE name = :name",
                        "parameters": {"name": node_name}
                    },
                    auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS)
                )

                if check_response.status_code == 200:
                    existing_nodes = check_response.json().get('result', [])
                    if existing_nodes:
                        logger.warning(f"Link with name '{node_name}' already exists")
                        return JsonResponse({'error': 'Node with this name already exists'}, status=409)

            # Проверка по содержанию
            check_response = requests.post(
                config_settings.ORIENT_COMMAND_URL,
                headers={'Content-Type': 'application/json'},
                json={
                    "command": "SELECT FROM link WHERE content = :content",
                    "parameters": {"content": node_content}
                },
                auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS)
            )

            if check_response.status_code == 200:
                existing_nodes = check_response.json().get('result', [])
                if existing_nodes:
                    logger.warning(f"Link with this content already exists")
                    return JsonResponse({'error': 'Node with this content already exists'}, status=409)

        # Создание параметризованного запроса
        create_command = {
            "command": f"CREATE VERTEX {node_class} SET content = :content",
            "parameters": {"content": node_content}
        }

        # Добавляем дополнительные параметры, если они есть
        if node_name:
            create_command["command"] += ", name = :name"
            create_command["parameters"]["name"] = node_name
        if node_uuid:
            create_command["command"] += ", uuid = :uuid"
            create_command["parameters"]["uuid"] = node_uuid

        # Создание узла
        response = requests.post(
            config_settings.ORIENT_COMMAND_URL,
            headers={'Content-Type': 'application/json'},
            json=create_command,
            auth=(config_settings.ORIENT_LOGIN, config_settings.ORIENT_PASS)
        )

        # Обработка ответа
        if response.status_code == 200:
            try:
                response_data = response.json()
                result = response_data.get('result', [])

                if result:
                    logger.info(f"Node created successfully: {node_class}")
                    log_action(f"Создание новой вершины {node_class} | {node_content}")
                    return JsonResponse({
                        'status': 'success',
                        'data': result[0] if isinstance(result, list) else result
                    }, status=201)
                else:
                    logger.error("Empty result in response")
                    return JsonResponse({'error': 'Node creation failed'}, status=500)

            except ValueError as e:
                logger.error(f"Error parsing JSON response: {e}")
                return JsonResponse({'error': 'Failed to parse response'}, status=500)
        else:
            logger.error(f"Database error: {response.status_code} - {response.text}")
            return JsonResponse({
                'error': f"Database operation failed",
                'details': response.text
            }, status=500)

    except json.JSONDecodeError:
        logger.error("Invalid JSON format in request")
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)
    except requests.RequestException as e:
        logger.error(f"Database connection error: {str(e)}")
        return JsonResponse({'error': 'Database connection error'}, status=500)
    except Exception as e:
        logger.exception(f"Unexpected error in create_node: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)